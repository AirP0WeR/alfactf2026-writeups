#!/usr/bin/env python3
"""Initial exploration of provision.xlsx — dump cell values/formulas per sheet."""
import openpyxl
from pathlib import Path

XLSX = "/Users/airp0wer/Projects/alfa-ctf-2026/tasks-2026/provision/artifacts/provision.xlsx"
OUT = Path("/Users/airp0wer/Projects/alfa-ctf-2026/tasks-2026/provision/artifacts")

wb = openpyxl.load_workbook(XLSX, data_only=False)
print("sheets:", wb.sheetnames)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"== {sn}: dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")

# Dump Panel formulas / values into a file
def dump(ws, path, formulas_only=False):
    with open(path, "w") as f:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if formulas_only and not (isinstance(v, str) and v.startswith("=")):
                    continue
                f.write(f"{c.coordinate}\t{type(v).__name__}\t{v!r}\n")

dump(wb["Panel"], OUT/"panel_all.tsv")
dump(wb["Panel"], OUT/"panel_formulas.tsv", formulas_only=True)
dump(wb["s1"], OUT/"s1_all.tsv")
dump(wb["s1"], OUT/"s1_formulas.tsv", formulas_only=True)
dump(wb["s2"], OUT/"s2_all.tsv")
dump(wb["s2"], OUT/"s2_formulas.tsv", formulas_only=True)
dump(wb["s3"], OUT/"s3_all.tsv")
dump(wb["s3"], OUT/"s3_formulas.tsv", formulas_only=True)

print("dumped")
